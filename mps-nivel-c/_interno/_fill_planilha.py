# -*- coding: utf-8 -*-
"""Preenche a PlanilhaIndicadores ASR (Nivel C).
- Col A: nome do documento / rótulo de seção
- Col B (CP): link para documentos organizacionais (valem para todos os projetos)
- Cols C-F (projetos): links para documentos específicos por projeto
- Linhas (T,L,P,N,NA): intocadas — preenchimento exclusivo do avaliador ASR
- Col B (Final) nas abas ORG: intocada
- AQU: mantido em branco (empresa não tem evidência de uso deste processo)
"""
import csv, re, openpyxl
from openpyxl.styles import Alignment, Font

SRC = '/root/.claude/uploads/5abbaf16-7c03-5c89-bd42-930844eb89cc/d820916e-PlanilhaIndicadores_SW_2024__NivelC.xlsx'
OUT = '/home/user/MPS-repository/mps-nivel-c/_interno/PlanilhaIndicadores_SW_2024_NivelC_PREENCHIDA.xlsx'
CSV = '/root/.claude/uploads/5abbaf16-7c03-5c89-bd42-930844eb89cc/216596c6-16.6_MAPADRIVE_IndicedeLinks.csv'

# ---- link index ----
LINKS = {}
with open(CSV, encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        doc = row['Documento'].strip(); tipo = row['Tipo'].strip().lower(); link = row['Link'].strip()
        if tipo == 'md': continue
        LINKS.setdefault(doc.split('_',1)[0], {})[tipo] = (doc, link)

# Entradas manuais — docs cujo nome no Drive não segue a convenção padrão
LINKS.setdefault('ATA-FRUKI01-008', {})['docx'] = ('ATA-FRUKI01-008', 'https://drive.google.com/file/d/11alVpwlnsPjgd1_PO4OzMBXeJLRz6PXi/view?usp=drivesdk')
LINKS.setdefault('REG-MED-001', {})['docx'] = ('REG-MED-001', 'https://drive.google.com/file/d/1FOAiyoGlyqXioyDAXV4gCyqsJZvaJ1aB/view?usp=drivesdk')

MISSING = set()

def get_url(code, ext=None):
    e = LINKS.get(code)
    if not e: MISSING.add(code); return None
    if ext and ext in e: return e[ext][1]
    if 'docx' in e: return e['docx'][1]
    return list(e.values())[0][1]

def get_fname(code, ext=None):
    e = LINKS.get(code)
    if not e: return code
    if ext and ext in e: return e[ext][0]
    if 'docx' in e: return e['docx'][0]
    return list(e.values())[0][0]

# ============ PROJECT ROLE MAPS ============
FRU = {
 'TAP':['TAP-FRUKI01-001','TAP-FRUKI01-002'],
 'PLA':['PLA-FRUKI01-001','PLA-FRUKI01-002'],
 'ADAP':['ADAP-FRUKI01-001','ADAP-FRUKI01-002'],
 'REQ':['REQ-FRUKI01-001','REQ-FRUKI01-002'],
 'RASTR':['RASTR-FRUKI01-001','RASTR-FRUKI01-002'],
 'PCP':['PCP-FRUKI01-001','PCP-FRUKI01-002'],
 'VV':['VV-FRUKI01-001','VV-FRUKI01-002'],
 'ITP':['ITP-FRUKI01-001','ITP-FRUKI01-002'],
 'GDE':['GDE-FRUKI01-001'],
 'GCO':['GCO-FRUKI01-001'],
 'MED':['MED-FRUKI01-001'],
 'GQA':['GQA-FRUKI01-001'],
 'RAC':['RAC-FRUKI01-001'],
 'CR':['CR-FRUKI01-001'],
 'LI':['LI-FRUKI01-001'],
 'TAE':['TAE-FRUKI01-001','TAE-FRUKI01-002'],
 'ATAK':['ATA-FRUKI01-001','ATA-FRUKI01-008'],
 'ATA_METAS':['ATA-FRUKI01-002'],
 'ATAA':['ATA-FRUKI01-003'],
 'ATA_VAL':['ATA-FRUKI01-004','ATA-FRUKI01-005','ATA-FRUKI01-006','ATA-FRUKI01-007'],
 'GEST':['GEST-FRUKI01-001'],
 'REV':[],'CTQ':[],'RELVV':[],'CAP':[],
}
AASPAP = {
 'TAP':['TAP-AASPAP01-001'],
 'PLA':['PLA-AASPAP01-001'],
 'ADAP':['ADAP-AASPAP01-001'],
 'REQ':['REQ-AASPAP01-001'],
 'RASTR':['RASTR-AASPAP01-001'],
 'PCP':['PCP-AASPAP01-001'],
 'REV':['REV-AASPAP01-001','REV-AASPAP01-002'],
 'VV':['VV-AASPAP01-001'],
 'RELVV':['REL-VV-AASPAP01-001'],
 'ITP':['ITP-AASPAP01-001'],
 'GCO':['GCO-AASPAP01-001'],
 'GDE':['GDE-AASPAP01-001'],
 'MED':['MED-AASPAP01-001'],
 'GQA':['GQA-AASPAP01-001'],
 'RAC':['RAC-AASPAP01-001'],
 'ATAK':['ATA-AASPAP01-001'],
 'ATAA':['ATA-AASPAP01-002','ATA-AASPAP01-003','ATA-AASPAP01-004','ATA-AASPAP01-005'],
 'ATA_METAS':['MQ-AASPAP01-001'],
 'CR':['CR-AASPAP01-001','CR-AASPAP01-002'],
 'GEST':['GEST-AASPAP01'],
 'LI':[],'CTQ':[],'CAP':[],'ATA_VAL':[],'TAE':[],
}
PROF = {
 'TAP':['TAP-PROFARMA01-001'],
 'PLA':['PLA-PROFARMA01-001'],
 'ADAP':['ADAP-PROFARMA01-001'],
 'REQ':['REQ-PROFARMA01-001'],
 'RASTR':['RASTR-PROFARMA01-001'],
 'PCP':['PCP-PROFARMA01-001'],
 'REV':['REV-PROFARMA01-001'],
 'VV':['VV-PROFARMA01-001'],
 'CTQ':['CTQ-PROFARMA01-001'],
 'RELVV':['REL-VV-PROFARMA01-001'],
 'ITP':['ITP-PROFARMA01-001'],
 'GCO':['GCO-PROFARMA01-001'],
 'GDE':['GDE-PROFARMA01-001'],
 'MED':['MED-PROFARMA01-001'],
 'GQA':['GQA-PROFARMA01-001'],
 'RAC':['RAC-PROFARMA01-001'],
 'ATAK':['ATA-PROFARMA01-001'],
 'ATA_VAL':['ATA-PROFARMA01-002'],
 'CR':['CR-PROFARMA01-001'],
 'LI':['LI-PROFARMA01-001'],
 'TAE':['TAE-PROFARMA01-001'],
 'GEST':['GEST-PROFARMA01'],
 'ATAA':[],'ATA_METAS':[],'CAP':[],
}
AASP = {
 'TAP':['TAP-AASPCNJ01-001'],
 'PLA':['PLA-AASPCNJ01-001'],
 'ADAP':['ADAP-AASPCNJ01-001'],
 'REQ':['REQ-AASPCNJ01-001'],
 'RASTR':['RASTR-AASPCNJ01-001'],
 'PCP':['PCP-AASPCNJ01-001'],
 'REV':['REV-AASPCNJ01-001'],
 'VV':['VV-AASPCNJ01-001'],
 'RELVV':['REL-VV-AASPCNJ01-001'],
 'ITP':['ITP-AASPCNJ01-001'],
 'GCO':['GCO-AASPCNJ01-001'],
 'GDE':['GDE-AASPCNJ01-001'],
 'MED':['MED-AASPCNJ01-001'],
 'GQA':['GQA-AASPCNJ01-001'],
 'RAC':['RAC-AASPCNJ01-001'],
 'ATAK':['ATA-AASPCNJ01-001'],
 'CR':['CR-AASPCNJ01-001'],
 'GEST':['GEST-AASPCNJ01-001'],
 'CAP':['CAP-AASPCNJ01-001'],
 'LI':[],'CTQ':[],'ATAA':[],'ATA_METAS':[],'ATA_VAL':[],'TAE':[],
}

PROJECTS = [('FRUKI',FRU),('AASP_AP',AASPAP),('PROFARMA',PROF),('AASP_CNJ',AASP)]
PCOL = {'FRUKI':3,'AASP_AP':4,'PROFARMA':5,'AASP_CNJ':6}  # C, D, E, F

PROJ_NAMES = {
 'FRUKI':   'FTFRUKI — Super App Força de Vendas',
 'AASP_AP': 'AASP — Andamento Processuais',
 'PROFARMA':'PROFARMA / D1000 — Cadastro de Clientes',
 'AASP_CNJ':'AASP — CNJ Integração',
}

ROLE_LABELS = {
 'TAP':'TAP — Termo de Abertura do Projeto',
 'PLA':'PLA — Plano de Projeto',
 'ADAP':'ADAP — Registro de Adaptacao do Processo',
 'REQ':'REQ — Documento de Requisitos',
 'RASTR':'RASTR — Matriz de Rastreabilidade',
 'PCP':'PCP — Documento de Design / Projeto e Construcao',
 'REV':'REV — Registro de Revisao por Pares',
 'VV':'VV — Plano de Verificacao e Validacao',
 'RELVV':'REL-VV — Relatorio de Execucao de Testes',
 'CTQ':'CTQ — Checklist de Qualidade',
 'GCO':'GCO — Registro de Configuracao',
 'GDE':'GDE — Registro de Analise de Decisao',
 'MED':'MED — Registro de Medicao',
 'GQA':'GQA — Registro de Auditoria (GQA)',
 'RAC':'RAC — Relatorio de Acompanhamento',
 'ATAK':'ATA — Ata de Kick-off / Reunioes de requisitos',
 'ATAA':'ATA — Atas de Acompanhamento de Projeto',
 'ATA_METAS':'ATA — Ata de Definicao de Metas de Qualidade',
 'ATA_VAL':'ATA — Atas de Validacao / Aceite',
 'CR':'CR — Change Request',
 'LI':'LI — Registro de Licoes Aprendidas',
 'TAE':'TAE — Termo de Aceite da Entrega',
 'GEST':'GEST — Planilha de Gestao do Projeto',
 'ITP':'ITP — Estrategia de Integracao do Produto',
 'CAP':'CAP — Registro de Capacitacao da Equipe',
}

IND_ROLES = {
 'GPR 1':['TAP','PLA'],'GPR 2+':['ADAP'],'GPR 3+':['PLA'],'GPR 4+':['PLA'],
 'GPR 5':['PLA','GEST'],'GPR 6':['PLA'],'GPR 7+':['PLA'],'GPR 8':['PLA'],'GPR 9':['PLA'],
 'GPR 10+':['PLA','GEST'],'GPR 11':['PLA'],'GPR 12+':['PLA'],'GPR 13+':['ATAK','PLA'],
 'GPR 14+':['RAC','GEST'],'GPR 15':['RAC'],'GPR 16':['RAC','TAE'],'GPR 17+':['RAC','GEST'],
 'GPR 18+':['RAC','CR'],'GPR 19+':['LI','RAC'],'GPR 20':['LI'],
 'REQ 1':['REQ','ATAK','ATA_METAS'],'REQ 2+':['REQ'],'REQ 3':['REQ','ATAK'],'REQ 4':['RASTR'],
 'REQ 5':['RASTR','GQA'],'REQ 6':['REQ'],'REQ 7':['REQ','ATAA','RELVV','CTQ'],
 'PCP 1+':['PCP'],'PCP 2':['REV','PCP'],'PCP 3+':['PCP'],
 'ITP 1+':['ITP'],'ITP 2':['ITP'],'ITP 3+':['ITP','REV'],'ITP 4':['ITP'],
 'ITP 5+':['ITP','VV','RELVV'],'ITP 6':['TAE','ATAA','ITP'],
 'VV 1':['VV'],'VV 2':['VV','REV'],'VV 3+':['VV','CTQ'],
 'VV 4':['RELVV','REV','GQA','ATA_VAL'],'VV 5':['RELVV','ATAA','ATA_VAL'],
}

# ============ ORG SHEETS ============
GCO_REG =[('GCO-FRUKI01-001',None,None),('GCO-AASPAP01-001',None,None),
          ('GCO-PROFARMA01-001',None,None),('GCO-AASPCNJ01-001',None,None)]
MED_REG =[('MED-FRUKI01-001',None,None),('MED-AASPAP01-001',None,None),
          ('MED-PROFARMA01-001',None,None),('MED-AASPCNJ01-001',None,None)]
GDE_REG =[('GDE-FRUKI01-001',None,None),('GDE-AASPAP01-001',None,None),
          ('GDE-PROFARMA01-001',None,None),('GDE-AASPCNJ01-001',None,None)]
GQA_REG =[('GQA-FRUKI01-001',None,None),('GQA-AASPAP01-001',None,None),
          ('GQA-PROFARMA01-001',None,None),('GQA-AASPCNJ01-001',None,None)]
LI_REG  =[('LI-FRUKI01-001',None,None),('LI-PROFARMA01-001',None,None)]
ADAP_REG=[('ADAP-FRUKI01-001',None,None),('ADAP-FRUKI01-002',None,None),
          ('ADAP-AASPAP01-001',None,None),('ADAP-PROFARMA01-001',None,None),
          ('ADAP-AASPCNJ01-001',None,None)]
REGCAP=[(f'REG-CAP-{n}',None,None) for n in ['001','001B','002','002B','003','004','005','006','007','008','009','010','011','012','013']]
AVACAP=[(f'AVA-CAP-{n}',None,None) for n in ['001','002','003','004','005']]

# AQU mantido em branco — sem evidência de uso (equipe própria, sem subcontratação)
ORG_EV = {
 'AQU 1':[],'AQU 2':[],'AQU 3+':[],'AQU 4+':[],'AQU 5':[],
 'GCO 1':[('PLA-GCO-001','§3',None),('GUIA-GCO-001',None,None),('CONV-ORG-001',None,None),('PRO-GCO-001',None,None)],
 'GCO 2':[('PLA-GCO-001','§4',None),('GUIA-GCO-001',None,None),('PRO-GCO-001',None,None)],
 'GCO 3':[('PLA-GCO-001','§5',None)]+GCO_REG+[('PRO-GCO-001',None,None)],
 'GCO 4':[('PLA-GCO-001','§6',None),('CONV-ORG-001',None,None)]+GCO_REG+[('PRO-GCO-001',None,None)],
 'GCO 5':[('PLA-GCO-001','§7',None),('EST-GPC-001',None,None)]+GQA_REG+[('PRO-GCO-001',None,None)],
 'MED 1':[('PLA-MED-001','§2',None),('PRO-MED-001',None,None)],
 'MED 2':[('PLA-MED-001','§3',None),('PRO-MED-001',None,None)],
 'MED 3+':[('PLA-MED-001','§4',None),('REG-MED-001',None,None)]+MED_REG+[('PRO-MED-001',None,None)],
 'MED 4+':[('PLA-MED-001','§5',None)]+MED_REG+[('PRO-MED-001',None,None)],
 'MED 5':[('PLA-MED-001','§6',None),('PRO-MED-001',None,None)],
 'MED 6':[('PLA-MED-001','§7',None),('REG-OSW-001',None,None),('PRO-MED-001',None,None)],
 'MED 7':[('PLA-MED-001','§8',None),('REG-MED-001',None,None),('PRO-MED-001',None,None)],
 'GDE 1':[('PRO-GDE-001','§3',None)],
 'GDE 2':[('PRO-GDE-001','§4',None)]+GDE_REG,
 'GDE 3':[('PRO-GDE-001','§4',None)]+GDE_REG,
 'GDE 4':[('PRO-GDE-001','§4',None),('TPL-GDE-001',None,None)]+GDE_REG,
 'GDE 5':[('PRO-GDE-001','§5',None)],
 'GDE 6':[('PRO-GDE-001','§5',None),('TPL-GDE-001',None,None)]+GDE_REG,
 'CAP 1+':[('PLA-CAP-001','§3',None),('PRO-CAP-001',None,None)],
 'CAP 2':[('PLA-CAP-001','§4',None),('TPL-CAP-001',None,None),('PRO-CAP-001',None,None)]+REGCAP,
 'CAP 3':[('PLA-CAP-001','§5',None),('REL-CAP-001',None,None),('PRO-CAP-001',None,None)]+AVACAP,
 'CAP 4':[('PLA-CAP-001','§6',None),('REG-CAP-CV-001',None,None),('PRO-CAP-001',None,None)],
 'GPC 1':[('PLA-GPC-001','§2',None)],
 'GPC 2+':[('PRO-GPC-001',None,None),('GUIA-GPC-001',None,None)],
 'GPC 3':[('EST-GPC-001',None,None),('TPL-GPC-001',None,None),('GQA-ORG-001',None,None)]+GQA_REG,
 'GPC 4+':[('PLA-GPC-001','§5.1',None),('REG-GPC-001',None,None)]+LI_REG,
 'GPC 5+':[('REG-GPC-002',None,None),('PLA-GPC-001','§5',None)],
 'GPC 6':[('PRO-GPC-002',None,None),('ATA-GPC-001',None,None),('REG-GPC-002',None,None)],
 'GPC 7':[('EST-GPC-002',None,None),('ATA-GPC-001',None,None)],
 'GPC 8':[('PLA-GPC-001','§3',None),('GUIA-GCO-001',None,None)],
 'GPC 9':[('PLA-MED-001','§4 e §8',None),('REG-MED-001',None,None)],
 'GPC 10':[('PLA-GPC-001','§4',None)]+ADAP_REG,
 'GPC 11':[('PLA-GPC-001','§6',None),('REG-GPC-002',None,None),('ATA-GPC-001',None,None)],
 'OSW 1':[('POL-ORG-001',None,None),('REG-OSW-002',None,None)],
 'OSW 2':[('PRO-OSW-001','§3',None),('PLA-CAP-001','§6.1',None)],
 'OSW 3':[('PRO-OSW-001','§6',None),('ATA-GPC-001',None,None),('REG-OSW-001',None,None)],
 'OSW 4+':[('PRO-OSW-001','§4',None),('MAPA-ORG-001',None,'docx')],
 'OSW 5+':[('PRO-OSW-001','§5',None),('EST-GPC-002',None,None)],
 'OSW 6':[('PRO-OSW-001','§6',None),('PLA-MED-001','§5',None)],
 'OSW 7':[('PRO-OSW-001','§7',None),('ATA-GPC-001',None,None),('REG-OSW-002',None,None)],
 'OSW 8':[('PRO-OSW-002','§3',None),('REG-OSW-001',None,None)],
 'OSW 9':[('PRO-OSW-002','§4',None),('REG-OSW-001',None,None)],
 'OSW 10':[('PRO-OSW-002','§5',None),('REG-OSW-001',None,None)],
}
ORG_NO_RATING = set()  # sem preenchimento de avaliação — exclusivo do avaliador

# ============ ESTILOS ============
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='top')
BLUE = Font(color='0563C1', underline='single')

def write_link(ws, row, col, url):
    cell = ws.cell(row=row, column=col, value='x')
    cell.hyperlink = url
    cell.font = BLUE
    cell.alignment = CENTER

def find_blocks(ws):
    blocks = []; cur = None
    for r in range(4, ws.max_row+1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a and isinstance(a, str):
            m = re.match(r'^([A-Z]{2,3}\s*\d+\+?)', a.strip())
            if m:
                code = re.sub(r'\s+', ' ', m.group(1)).strip()
                cur = [code, r, None]; blocks.append(cur)
        if b and isinstance(b, str) and b.startswith('(T,L,P,N,NA') and cur and cur[2] is None:
            cur[2] = r
    return blocks

def blank_rows_for(ws, irow, lrow):
    end = lrow if lrow else irow + 10
    return [r for r in range(irow + 1, end) if ws.cell(row=r, column=1).value is None]

# ============ WRITE ============
wb = openpyxl.load_workbook(SRC)
pmap_dict = dict(PROJECTS)

# ---- cabeçalho de projeto (R3) nas abas de processo ----
for sheet in ['GPR','REQ','PCP','ITP','VV']:
    ws = wb[sheet]
    for pname, col in PCOL.items():
        ws.cell(row=3, column=col, value=PROJ_NAMES.get(pname, pname)).alignment = CENTER

# ---- abas de projeto (GPR, REQ, PCP, ITP, VV) ----
for sheet in ['GPR','REQ','PCP','ITP','VV']:
    ws = wb[sheet]
    for code, irow, lrow in find_blocks(ws):
        if code not in IND_ROLES: continue
        roles = IND_ROLES[code]
        blanks = blank_rows_for(ws, irow, lrow)
        row_cursor = 0

        for role in roles:
            label = ROLE_LABELS.get(role, role)
            proj_urls = {}
            for pname, pmap in PROJECTS:
                for c in pmap.get(role, []):
                    url = get_url(c)
                    if url: proj_urls[pname] = url; break

            if not proj_urls: continue
            if row_cursor >= len(blanks): break

            r = blanks[row_cursor]; row_cursor += 1
            ws.cell(row=r, column=1, value=label).alignment = Alignment(vertical='top')
            for pname, col in PCOL.items():
                if pname in proj_urls:
                    write_link(ws, r, col, proj_urls[pname])
        # linha (T,L,P,N,NA) — não preencher

# ---- abas organizacionais (GCO, MED, GDE, CAP, GPC, OSW) — AQU em branco ----
for sheet in ['AQU','GCO','MED','GDE','CAP','GPC','OSW']:
    ws = wb[sheet]
    for code, irow, lrow in find_blocks(ws):
        if code not in ORG_EV: continue
        items = ORG_EV[code]
        if not items: continue  # AQU e qualquer indicador sem evidência = em branco
        blanks = blank_rows_for(ws, irow, lrow)
        row_cursor = 0

        for item in items:
            if row_cursor >= len(blanks): break
            if isinstance(item, str):
                r = blanks[row_cursor]; row_cursor += 1
                ws.cell(row=r, column=1, value=item).alignment = WRAP
            else:
                doc_code, sec, ext = item
                url = get_url(doc_code, ext)
                if not url: continue
                fname = get_fname(doc_code, ext)
                label = f"{fname} {sec}" if sec else fname
                r = blanks[row_cursor]; row_cursor += 1
                ws.cell(row=r, column=1, value=label).alignment = Alignment(vertical='top')
                write_link(ws, r, 3, url)  # col C = Fonte da Evidência (ORG)
        # linha (T,L,P,N,NA) — não preencher

# ============ CP_Projeto ============
# Estrutura do template (nosso d820916e):
# R3: cabeçalho de projeto; R4:(i) R5-8 blank; R9:(ii) R10-13 blank;
# R14:(iii) R15-18 blank; R19:(iv) R20-23 blank; R24:(v) R25-28 blank;
# R29:(vi) R30-32 blank; R33:(vii) R34-36 blank; R37:(T,L,P,N,NA) — não tocar
#
# Convenção: col B = evidência organizacional (vale todos os projetos)
#            cols C-F = evidência específica por projeto

ws_cp = wb['CP_Projeto']

# R3: nomes dos projetos
for pname, col in PCOL.items():
    ws_cp.cell(row=3, column=col, value=PROJ_NAMES[pname]).alignment = CENTER

def cp_org_link(row, col_b, code, ext=None, sec=None):
    url = get_url(code, ext)
    if not url: return False
    fname = get_fname(code, ext)
    label = f"{fname} {sec}" if sec else fname
    ws_cp.cell(row=row, column=1, value=label).alignment = Alignment(vertical='top')
    write_link(ws_cp, row, col_b, url)
    return True

def cp_proj_links(row, role):
    written = False
    for pname, pmap in PROJECTS:
        col = PCOL[pname]
        for c in pmap.get(role, []):
            url = get_url(c)
            if url:
                write_link(ws_cp, row, col, url)
                written = True
                break
    return written

# (i) R4 — texto informativo
ws_cp.cell(row=5, column=1,
    value='Todas as evidências de execução dos resultados dos processos nos projetos (GPR, REQ, PCP, ITP, VV)'
).alignment = WRAP

# (ii) R9 — processo padrão e adaptação
cp_org_link(10, 2, 'PRO-GPR-001')
cp_org_link(11, 2, 'GUIA-GPC-001')
ws_cp.cell(row=12, column=1, value='ADAP — Registro de Adaptacao do Processo').alignment = Alignment(vertical='top')
cp_proj_links(12, 'ADAP')
cp_org_link(13, 2, 'MAPA-ORG-001', ext='docx')

# (iii) R14 — pessoas preparadas
cp_org_link(15, 2, 'MAPA-ORG-001', ext='docx')
cp_org_link(16, 2, 'REG-CAP-CV-001')
cp_org_link(17, 2, 'PLA-CAP-001')

# (iv) R19 — verificação objetiva do processo
ws_cp.cell(row=20, column=1, value='GEST — Planilha de Gestao do Projeto').alignment = Alignment(vertical='top')
cp_proj_links(20, 'GEST')
ws_cp.cell(row=21, column=1, value='GQA — Registro de Auditoria (GQA)').alignment = Alignment(vertical='top')
cp_proj_links(21, 'GQA')

# (v) R24 — produtos de trabalho avaliados
ws_cp.cell(row=25, column=1, value='GEST — Planilha de Gestao do Projeto').alignment = Alignment(vertical='top')
cp_proj_links(25, 'GEST')
ws_cp.cell(row=26, column=1, value='GQA — Registro de Auditoria (GQA)').alignment = Alignment(vertical='top')
cp_proj_links(26, 'GQA')

# (vi) R29 — oportunidades de melhoria
ws_cp.cell(row=30, column=1, value='LI — Registro de Licoes Aprendidas').alignment = Alignment(vertical='top')
cp_proj_links(30, 'LI')
cp_org_link(31, 2, 'REG-GPC-001')

# (vii) R33 — informações e ativos de processo
cp_org_link(34, 2, 'PLA-GPC-001')
cp_org_link(35, 2, 'REG-GPC-002')
# R37 = (T,L,P,N,NA) — não tocar

# ============ CP_Organizacional ============
# Estrutura do template:
# R2: cabeçalhos de processo (AQU=C2, GCO=C4, MED=C6, GDE=C8, CAP=C10, GPC=C12, OSW=C14)
# R4: seção E/D/C; R5:(i) R6-9 blank; R10:(ii) R11-14 blank; R15:(iii) R16-19 blank;
# R20:(iv) R21-24 blank; R25:(v) R26-29 blank; R30:(vi) R31-34 blank;
# R35:(vii) R36-39 blank; R40:(T,L,P,N,NA) — não tocar
#
# Cols de evidência: AQU=2(branco), GCO=4, MED=6, GDE=8, CAP=10, GPC=12, OSW=14

ws_cpo = wb['CP_Organizacional']

# Colunas por processo (col de Fonte da Evidência)
ORG_COLS = {'GCO':4,'MED':6,'GDE':8,'CAP':10,'GPC':12,'OSW':14}
ALL_ORG_COLS = list(ORG_COLS.values())  # [4,6,8,10,12,14]

def cpo_link(row, cols, code, ext=None, sec=None, label_override=None):
    url = get_url(code, ext)
    if not url: return False
    fname = get_fname(code, ext)
    label = label_override or (f"{fname} {sec}" if sec else fname)
    ws_cpo.cell(row=row, column=1, value=label).alignment = Alignment(vertical='top')
    for col in (cols if isinstance(cols, list) else [cols]):
        write_link(ws_cpo, row, col, url)
    return True

# (i) R5 — texto informativo
ws_cpo.cell(row=6, column=1,
    value='Todas as evidências de execução dos resultados — ver abas GCO, MED, GDE, CAP, GPC, OSW'
).alignment = WRAP

# (ii) R10 — processo padrão
# linha 11: um PRO por processo
ws_cpo.cell(row=11, column=1, value='Definição de processos (PRO-*)').alignment = Alignment(vertical='top')
for proc, col in ORG_COLS.items():
    code_map = {'GCO':'PRO-GCO-001','MED':'PRO-MED-001','GDE':'PRO-GDE-001',
                'CAP':'PRO-CAP-001','GPC':'PRO-GPC-001','OSW':'PRO-OSW-001'}
    url = get_url(code_map[proc])
    if url: write_link(ws_cpo, 11, col, url)
# linha 12: complementos GPC e OSW
ws_cpo.cell(row=12, column=1, value='Definição de processos (complemento GPC/OSW)').alignment = Alignment(vertical='top')
for code, col in [('PRO-GPC-002',12),('PRO-OSW-002',14),('GUIA-GPC-001',12)]:
    url = get_url(code)
    if url: write_link(ws_cpo, 12, col, url)
# linha 13: MAPA-ORG-001 para todos
cpo_link(13, ALL_ORG_COLS, 'MAPA-ORG-001', ext='docx',
         label_override='MAPA-ORG-001 — Matriz de Papeis e Responsabilidades')

# (iii) R15 — pessoas preparadas
cpo_link(16, ALL_ORG_COLS, 'MAPA-ORG-001', ext='docx',
         label_override='MAPA-ORG-001 — Matriz de Papeis e Responsabilidades')
cpo_link(17, ALL_ORG_COLS, 'REG-CAP-CV-001',
         label_override='REG-CAP-CV-001 — Curriculos e Evidencias de Competencia')
cpo_link(18, ALL_ORG_COLS, 'PLA-CAP-001')

# (iv) R20 — verificação objetiva
cpo_link(21, ALL_ORG_COLS, 'GQA-ORG-001',
         label_override='GQA-ORG-001 — Auditoria Organizacional')
cpo_link(22, [12], 'EST-GPC-001',
         label_override='EST-GPC-001 — Estrategia de Auditoria de GQA')

# (v) R25 — produtos de trabalho avaliados
cpo_link(26, ALL_ORG_COLS, 'GQA-ORG-001',
         label_override='GQA-ORG-001 — Auditoria Organizacional')
cpo_link(27, [12], 'TPL-GPC-001',
         label_override='TPL-GPC-001 — Template de Auditoria de GQA')

# (vi) R30 — oportunidades de melhoria
cpo_link(31, ALL_ORG_COLS, 'PLA-GPC-001',
         label_override='PLA-GPC-001 — Plano de Melhoria de Processos')
cpo_link(32, [12], 'REG-GPC-001',
         label_override='REG-GPC-001 — Registro de Licoes Aprendidas Organizacionais')

# (vii) R35 — informações e ativos de processo
cpo_link(36, [12], 'REG-GPC-002',
         label_override='REG-GPC-002 — Registro de Dados de Desempenho de Processos')
cpo_link(37, ALL_ORG_COLS, 'PLA-GPC-001',
         label_override='PLA-GPC-001 — Plano de Melhoria de Processos')
# R40 = (T,L,P,N,NA) — não tocar

wb.save(OUT)
print("Saved:", OUT)
print(f"\nMISSING CODES ({len(MISSING)}):")
for m in sorted(MISSING): print("  -", m)
