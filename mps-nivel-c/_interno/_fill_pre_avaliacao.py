"""
Preenche PlanilhaAvaliacao_TIMEWARE_SW_C_PREENCHIDA.xlsx
Abas: PROJETOS, PLANO (col A), EQUIPE, EQUIPE_ORG
Dados extraídos do repositório (TAPs, MAPA-ORG-001, CAP, ATAs).
"""

import shutil
import datetime
from openpyxl import load_workbook

SRC = '/root/.claude/uploads/5abbaf16-7c03-5c89-bd42-930844eb89cc/97119115-planilha_1_pre_avali_Avaliac_a_o_2024_TIMEWARE_SW_C.xlsx'
OUT = '/home/user/MPS-repository/mps-nivel-c/_interno/PlanilhaAvaliacao_TIMEWARE_SW_C_PREENCHIDA.xlsx'

shutil.copy(SRC, OUT)
wb = load_workbook(OUT)


# ─────────────────────────────────────────
# PROJETOS
# Headers (R4): Nome | Descrição | Tipo | Cliente | Ciclo de vida/Sprint | Esforço horas |
#               Total envolvidos | Gerente | Data início | Data término |
#               Fase | Importância | Seleção | Justificativa
# ─────────────────────────────────────────
ws = wb['PROJETOS']

# R2:A2:B2 merged = label; valor vai em C2 (col 3)
# R3:A3:B3 merged = label; valor vai em C3 (col 3)
ws.cell(2, 3).value = 8   # total projetos na UO
ws.cell(3, 3).value = 1   # número de GPs

projects = [
    dict(
        nome='FTFRUKI — Super App Força de Vendas',
        descricao=(
            'Desenvolvimento de módulos do SuperApp Fruki (aplicativo mobile de força de vendas '
            'em React Native): Pacote 1 — módulo de Metas e Remuneração Variável; '
            'Pacote Final 24 — módulos de Pedidos Não Alocados, Regra de Ouro e PDV.'
        ),
        tipo='Evolução Produto',
        cliente='Fruki Bebidas S.A.',
        ciclo='Concluído',
        esforco=2058,   # realizado total (~847h Pacote 1 + ~1.211h Pacote Final 24; MED-FRUKI01-001 v1.1)
        envolvidos=5,
        gerente='Abraão Oliveira',
        inicio=datetime.date(2025, 6, 5),
        termino=datetime.date(2026, 1, 15),
        fase='Concluído',
        sprint_ou_fase='Concluído',
        importancia='Alta',
        selecao='Sim',
        justificativa=(
            'Projeto com conjunto completo de evidências MPS; duas fases documentadas '
            'com TAPs, planos, atas de acompanhamento, relatórios de V&V e lições aprendidas.'
        ),
    ),
    dict(
        nome='AASP — Andamento Processuais',
        descricao=(
            'Refatoração da solução legada de captura de andamentos processuais para '
            'arquitetura DataJud/CNJ + WorkerAndamentos, com integração EPROC/ESAJ, '
            'RabbitMQ e Elasticsearch.'
        ),
        tipo='Evolução Produto',
        cliente='AASP — Associação dos Advogados de São Paulo',
        ciclo='Construção',
        esforco=362,   # realizado Fases 1–4 (GEST-AASPAP01); Fase 5 implantação em apuração
        envolvidos=5,
        gerente='Abraão Oliveira',
        inicio=datetime.date(2025, 12, 15),
        termino=None,
        fase='Construção',
        sprint_ou_fase='Fase 5 — Implantação',
        importancia='Alta',
        selecao='Sim',
        justificativa=(
            'Projeto com evidências de todas as fases do processo-padrão; '
            'refatoração arquitetural documentada em quatro fases com atas de acompanhamento.'
        ),
    ),
    dict(
        nome='PROFARMA / D1000 — Cadastro de Clientes',
        descricao=(
            'Desenvolvimento do novo sistema de Cadastro de Clientes da Rede D1000 em .NET/Azure '
            '(Clean Architecture + PostgreSQL), substituindo base legada ITEC. '
            'Integração com VTEX, BlueSoft, Propz CRM e PBM; piloto na loja 9.'
        ),
        tipo='Desenvolvimento',
        cliente='Profarma S.A. / Rede D1000',
        ciclo='Concluído',
        esforco=5789,  # estimado PLA-PROFARMA01-001 §4 (orçamento de horas por papel)
        envolvidos=9,
        gerente='Abraão Oliveira',
        inicio=datetime.date(2025, 3, 17),
        termino=datetime.date(2026, 1, 29),
        fase='Concluído',
        sprint_ou_fase='Concluído',
        importancia='Alta',
        selecao='Sim',
        justificativa=(
            'Maior projeto em complexidade técnica e tamanho de equipe; '
            'ciclo de vida completo com evidências abrangentes de todos os processos MPS-SW Nível C.'
        ),
    ),
    dict(
        nome='AASP — CNJ Integração',
        descricao=(
            'Desenvolvimento de agente de captura de andamentos processuais via API DataJud/CNJ '
            '(WorkerAndamentos), consumindo RabbitMQ e consolidando histórico no Elasticsearch; '
            'integração com EPROC/ESAJ.'
        ),
        tipo='Desenvolvimento',
        cliente='AASP — Associação dos Advogados de São Paulo',
        ciclo='Construção',
        esforco=624,   # realizado Fases 1–6 parcial (GEST-AASPCNJ01 aba Medição)
        envolvidos=7,
        gerente='Abraão Oliveira',
        inicio=datetime.date(2025, 10, 1),
        termino=datetime.date(2026, 6, 30),
        fase='Construção',
        sprint_ou_fase='Fase 7 — Homologação / Encerramento',
        importancia='Alta',
        selecao='Sim',
        justificativa=(
            'Projeto com evidências de todos os processos MPS; contexto de implantação em produção '
            'documenta o ciclo completo de desenvolvimento até entrega.'
        ),
    ),
]

for i, proj in enumerate(projects):
    r = 5 + i
    ws.cell(r, 1).value = proj['nome']
    ws.cell(r, 2).value = proj['descricao']
    ws.cell(r, 3).value = proj['tipo']
    ws.cell(r, 4).value = proj['cliente']
    ws.cell(r, 5).value = proj['ciclo']
    if proj['esforco'] is not None:
        ws.cell(r, 6).value = proj['esforco']
    ws.cell(r, 7).value = proj['envolvidos']
    ws.cell(r, 8).value = proj['gerente']
    if proj['inicio']:
        ws.cell(r, 9).value = proj['inicio']
    if proj['termino']:
        ws.cell(r, 10).value = proj['termino']
    ws.cell(r, 11).value = proj['fase']
    ws.cell(r, 12).value = proj['importancia']
    ws.cell(r, 13).value = proj['selecao']
    ws.cell(r, 14).value = proj['justificativa']


# ─────────────────────────────────────────
# PLANO — nomes dos projetos em A17:A20
# As fórmulas do PLANO fazem vlookup em InstanciasSelecionadas (PARAMETROS)
# que é gerado dinamicamente a partir de PROJETOS onde Seleção='Sim'.
# Escrever o nome exato do projeto em col A ativa o lookup ao abrir no Excel.
# ─────────────────────────────────────────
ws2 = wb['PLANO']
for i, proj in enumerate(projects):
    ws2.cell(17 + i, 1).value = proj['nome']
    ws2.cell(17 + i, 4).value = proj['sprint_ou_fase']  # col D = Sprint ou Fase (manual)


# ─────────────────────────────────────────
# EQUIPE
# Colunas: C1=Seleção/Entrevista | C2=Nome | C3=Projeto | C4=Papel no projeto |
#          C5=Cargo organizacional | C6=Nome da Chefia
# Dados de TAP-FRUKI01-001/002, ATA-AASPAP01-002, TAP-PROFARMA01-001, CAP-AASPCNJ01-001
# ─────────────────────────────────────────
ws3 = wb['EQUIPE']

# Tupla: (seleção, nome, projeto, papel_no_projeto, cargo_org, chefia)
equipe_rows = [
    # ── FRUKI (TAP-FRUKI01-001 e TAP-FRUKI01-002) ──────────────────────
    ('Sim', 'Abraão Oliveira',          'FTFRUKI — Super App Força de Vendas', 'Gerente de Projeto / Product Owner',     'Gerente de Projetos',      'Wilson Yamada'),
    ('Não', 'Luca Watson',              'FTFRUKI — Super App Força de Vendas', 'Desenvolvedor Front-End (React Native)', 'Desenvolvedor',            'Abraão Oliveira'),
    ('Não', 'Thiago Gomes',             'FTFRUKI — Super App Força de Vendas', 'Desenvolvedor Front-End (React Native)', 'Desenvolvedor',            'Abraão Oliveira'),
    ('Não', 'Brenda Chrystie',          'FTFRUKI — Super App Força de Vendas', 'UX/UI Designer / Analista de Negócio',  'Designer UX/UI',           'Abraão Oliveira'),
    ('Não', 'Tiago Barbosa Nascimento', 'FTFRUKI — Super App Força de Vendas', 'COO / Gestão Comercial',                'CEO / Founder',            ''),
    # ── AASP_AP (ATA-AASPAP01-002, MQ-AASPAP01-001) ────────────────────
    ('Sim', 'Abraão Oliveira',            'AASP — Andamento Processuais', 'Gerente de Projeto',              'Gerente de Projetos',   'Wilson Yamada'),
    ('Sim', 'Cézar Hiraki Velázquez',     'AASP — Andamento Processuais', 'Tech Lead / Arquiteto / DevOps',  'Arquiteto de Software', 'Abraão Oliveira'),
    ('Não', 'Raony Chagas',               'AASP — Andamento Processuais', 'Desenvolvedor Sênior',            'Desenvolvedor Sênior',  'Cézar Hiraki Velázquez'),
    ('Não', 'Mateus Veloso',              'AASP — Andamento Processuais', 'Desenvolvedor (Suporte)',          'Desenvolvedor',         'Cézar Hiraki Velázquez'),
    ('Não', 'Caroline Sousa',             'AASP — Andamento Processuais', 'Analista de QA',                  'Analista de QA',        'Abraão Oliveira'),
    # ── PROFARMA (TAP-PROFARMA01-001) ───────────────────────────────────
    ('Sim', 'Abraão Oliveira',          'PROFARMA / D1000 — Cadastro de Clientes', 'Gerente de Projeto / Account Manager', 'Gerente de Projetos',   'Wilson Yamada'),
    ('Sim', 'Tiago Barbosa Nascimento', 'PROFARMA / D1000 — Cadastro de Clientes', 'Tech Lead',                            'CEO / Founder',         ''),
    ('Não', 'Erick Coelho',             'PROFARMA / D1000 — Cadastro de Clientes', 'Dev Principal / Arquiteto da solução', 'Arquiteto de Software', 'Tiago Barbosa Nascimento'),
    ('Não', 'Gustavo Mathias',          'PROFARMA / D1000 — Cadastro de Clientes', 'Dev Backend',                          'Desenvolvedor Sênior',  'Erick Coelho'),
    ('Não', 'Renan Kiyoshi',            'PROFARMA / D1000 — Cadastro de Clientes', 'Dev Backend',                          'Desenvolvedor Sênior',  'Erick Coelho'),
    ('Não', 'Cézar Hiraki Velázquez',   'PROFARMA / D1000 — Cadastro de Clientes', 'Dev Backend',                          'Arquiteto de Software', 'Tiago Barbosa Nascimento'),
    ('Não', 'João Cruz',                'PROFARMA / D1000 — Cadastro de Clientes', 'Dev Backend',                          'Desenvolvedor',         'Erick Coelho'),
    ('Não', 'David Buena',              'PROFARMA / D1000 — Cadastro de Clientes', 'Infra / DevOps',                       'Engenheiro DevOps',     'Cézar Hiraki Velázquez'),
    ('Não', 'Lucas Batista',            'PROFARMA / D1000 — Cadastro de Clientes', 'QA / Testes Automatizados',            'Analista de QA',        'Abraão Oliveira'),
    # ── AASP_CNJ (TAP-AASPCNJ01-001) ───────────────────────────────────
    ('Sim', 'Abraão Oliveira',        'AASP — CNJ Integração', 'Gerente de Projeto',               'Gerente de Projetos',       'Wilson Yamada'),
    ('Sim', 'Cézar Hiraki Velázquez', 'AASP — CNJ Integração', 'Tech Lead / DevOps / Arquiteto',   'Arquiteto de Software',     'Abraão Oliveira'),
    ('Não', 'Raony Chagas',           'AASP — CNJ Integração', 'Desenvolvedor Sênior (Principal)', 'Desenvolvedor Sênior',      'Cézar Hiraki Velázquez'),
    ('Não', 'Mateus Veloso',          'AASP — CNJ Integração', 'Desenvolvedor',                    'Desenvolvedor',             'Cézar Hiraki Velázquez'),
    ('Não', 'Lucas Batista',          'AASP — CNJ Integração', 'Desenvolvedor',                    'Desenvolvedor',             'Cézar Hiraki Velázquez'),
    ('Não', 'Caroline Sousa',         'AASP — CNJ Integração', 'Analista de QA',                   'Analista de QA',            'Abraão Oliveira'),
    ('Não', 'Jonathan Barbosa',       'AASP — CNJ Integração', 'Auditor GQA (independente)',        'Analista de Qualidade',     'Abraão Oliveira'),
]

for i, row_data in enumerate(equipe_rows):
    r = 4 + i
    sel, nome, proj, papel, cargo, chefia = row_data
    ws3.cell(r, 1).value = sel
    ws3.cell(r, 2).value = nome
    ws3.cell(r, 3).value = proj
    ws3.cell(r, 4).value = papel
    ws3.cell(r, 5).value = cargo
    ws3.cell(r, 6).value = chefia


# ─────────────────────────────────────────
# EQUIPE_ORG
# Colunas: C1=Seleção/Entrevista | C2=Nome | C3=Processo organizacional da UO |
#          C4=Papel no processo da UO | C5=Cargo organizacional | C6=Nome da Chefia
# Dados de MAPA-ORG-001_Matriz-de-Papeis-e-Responsabilidades.md (Seção 5)
# ─────────────────────────────────────────
ws4 = wb['EQUIPE_ORG']

# Tupla: (seleção, nome, processo_org, papel_no_processo, cargo_org, chefia)
equipe_org_rows = [
    ('Sim', 'Wilson Yamada',
     'OSW — Gerência Organizacional de Software',
     'COO / Responsável OSW e Portfólio',
     'COO (Chief Operating Officer)',
     'Tiago Barbosa Nascimento'),

    ('Sim', 'Silvio Baroni',
     'GPC — Gerência de Processos',
     'Coordenador do Time de Melhoria Contínua (SEPG)',
     'Gerente de Produto Sênior',
     'Wilson Yamada'),

    ('Sim', 'Silvio Baroni',
     'MED — Medição',
     'Responsável de Medição',
     'Gerente de Produto Sênior',
     'Wilson Yamada'),

    ('Sim', 'Abraão Oliveira',
     'GPC — Gerência de Processos',
     'Membro do Time de Melhoria Contínua (SEPG)',
     'Gerente de Projetos',
     'Wilson Yamada'),

    ('Sim', 'Caroline Jenifer',
     'GPC — Gerência de Processos',
     'GQA — Garantia da Qualidade de Processo (independente)',
     'Analista de QA Sênior',
     'Abraão Oliveira'),

    ('Sim', 'Flávio Fernandes',
     'GCO — Gerência de Configuração',
     'DevOps / GCO Baseline e Auditoria',
     'Engenheiro DevOps',
     'Cézar Hiraki Velázquez'),

    ('Sim', 'Guilherme Gomes',
     'CAP — Capacitação',
     'Responsável de RH / Capacitação',
     'Analista de RH',
     'Wilson Yamada'),

    ('Não', 'Klayton Roberto',
     'CAP — Capacitação',
     'Apoio RH / Capacitação',
     'Analista de RH',
     'Guilherme Gomes'),

    ('Não', 'Patricia Lima',
     'GPC — Gerência de Processos',
     'Membro do Time de Melhoria Contínua (SEPG)',
     'Analista de Processos',
     'Silvio Baroni'),

    ('Não', 'Mariana Teixeira',
     'GPC — Gerência de Processos',
     'Membro do Time de Melhoria Contínua (SEPG)',
     'Analista de Processos',
     'Silvio Baroni'),
]

for i, row_data in enumerate(equipe_org_rows):
    r = 4 + i
    sel, nome, processo, papel, cargo, chefia = row_data
    ws4.cell(r, 1).value = sel
    ws4.cell(r, 2).value = nome
    ws4.cell(r, 3).value = processo
    ws4.cell(r, 4).value = papel
    ws4.cell(r, 5).value = cargo
    ws4.cell(r, 6).value = chefia


wb.save(OUT)
print(f'Salvo: {OUT}')

# Verificação rápida
wb2 = load_workbook(OUT)
ws_p = wb2['PROJETOS']
print(f'\nPROJETOS R2B={ws_p["B2"].value}, R3B={ws_p["B3"].value}')
for r in range(5, 9):
    print(f'  R{r}: {ws_p.cell(r,1).value} | {ws_p.cell(r,11).value} | Seleção={ws_p.cell(r,13).value}')

ws_pl = wb2['PLANO']
print(f'\nPLANO A17-A20:')
for r in range(17, 21):
    print(f'  R{r}: {ws_pl.cell(r,1).value}')

ws_e = wb2['EQUIPE']
print(f'\nEQUIE (primeiras 5 linhas de dados):')
for r in range(4, 9):
    print(f'  R{r}: {ws_e.cell(r,2).value} | {ws_e.cell(r,3).value} | {ws_e.cell(r,4).value}')

ws_eo = wb2['EQUIPE_ORG']
print(f'\nEQUIE_ORG (primeiras 5 linhas):')
for r in range(4, 9):
    print(f'  R{r}: {ws_eo.cell(r,2).value} | {ws_eo.cell(r,3).value}')
