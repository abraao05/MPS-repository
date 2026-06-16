#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera o mapa de links do Google Drive (acervo oficial Timeware) para a planilha
de indicadores da ASR.

Saida: MAPA-DRIVE_Indice-de-Links.csv  (colunas: Pasta | Documento | Tipo | Link)
       MAPA-DRIVE_Indice-de-Links.xlsx (se openpyxl estiver disponivel)

Os dados (pasta, nome do arquivo, id) foram coletados via search_files do
conector Google Drive, pasta por pasta (parentId), em 13/06/2026.
O link e' montado pelo template oficial do Drive a partir do id.
"""
import csv
import os
import sys

LINK_TPL = "https://drive.google.com/file/d/{id}/view?usp=drivesdk"

# Cada linha: pasta ||| nome_do_arquivo_com_extensao ||| id_do_drive
DATA = r"""
(raiz)|||Avaliacão_Oficial_Planilha-Indicadores_SW_2024-Nivel C|||1iiroC88KQ1SPTis5VBGRxBChmZUizI1hubZXxpyzNBE
(raiz)|||Ata-de-criação-SEPG.pdf|||1EFI0RcRsilCo8p7NDsd71-f3bWtlWZMc
(raiz)|||TPL-GPC-002_Ata-Reuniao-SEPG.xlsx|||1sGkGUkG4K6JYd8AWGganOiN7E2bJZj8j
00_governanca|||CONV-ORG-001_Convencao-de-Nomenclatura-e-Versionamento.docx|||1Ag4-jg-EEHXvggj27dmSr_07uzkYh77k
00_governanca|||REG-GPC-001_Registro-de-Melhorias-de-Processo.docx|||1h_fkAu9MjScpjg3l9r3OHepLf0zU8eA9
00_governanca|||EST-GPC-002_Estrategia-de-Gerencia-de-Riscos.docx|||1jNY12kyO2w-DJKhz8aizdHI9Q7HBo3Pl
00_governanca|||ATA-GPC-001_Reuniao-Analise-Critica-de-Processos.docx|||1mSEGlpnnXBRwaVqlnsioWGVZQCfdyXVD
00_governanca|||PLA-GPC-001_Plano-de-Gestao-e-Melhoria-de-Processos.docx|||131GoQo25d8-YX5ejvh1DAAUeMWl1Mk-K
00_governanca|||PRO-OSW-001_Governanca-Organizacional-de-Processos.docx|||1aLxm5g-o_qTP6V_LSQybcTC7nzdMgUAt
00_governanca|||GQA-ORG-001_Auditoria-Organizacional.docx|||1vNp-iv47mRXMyLEeR3f1pEDN4JXA1Zhv
00_governanca|||REG-GPC-002_Plano-de-Implementacao-de-Melhorias.docx|||1GCLn8iMOKMAUf2w_FjpOSx40cmfO741E
00_governanca|||PRO-GPC-001_Processo-Padrao-Organizacional.docx|||1NU3XFUjMKuLcBIEW6DrgKIz0BNWi3KxV
00_governanca|||REG-OSW-001_Painel-de-Portfolio.docx|||1kAjVXkatpQad4EuU19Il6BzL0c8IkITD
00_governanca|||REG-OSW-002_Comunicacao-da-Politica-Organizacional.docx|||1nqnyQ3vWpXA8HPeQMRc2X0XuZNOUx1uu
00_governanca|||GUIA-GPC-001_Guia-de-Adaptacao-do-Processo-Padrao.docx|||1ha36PVPtCC4YIeBXc0qZ_aLGJVNxY3I0
00_governanca|||PRO-GPC-002_Definicao-do-Time-de-Melhoria-Continua.docx|||1lh--xCFmICNqD7GKkDmQ1F3MOPTgox1V
00_governanca|||POL-ORG-001_Politica-Organizacional-de-Processos.docx|||1GmI7A4qb2yBY7y6tfUmd7UEtsKXqSFhy
00_governanca|||PRO-OSW-002_Gestao-de-Portfolio-de-Projetos.docx|||1bZr4oVRjUSGwYvRuL4v_cEioy6gOgJro
00_governanca|||EST-GPC-001_Estrategia-de-Garantia-da-Qualidade.docx|||1PdyKvi2sk8adOJE2mKW3L5QumD9shj-v
00_governanca|||REG-OSW-003_Riscos-Organizacionais.xlsx|||1P2hhQaWTAbOlyifwPrbaEdF3kTwqLMFP
00_governanca|||Timeware_GQA_-_CHK-Auditoria_QA_do_QA_Jun2026.xlsx|||1JjU5851rE_9gVpEN9ySdnLYAiGseVaME
00_governanca|||GUIA-GPC-002_Guia-de-Auditorias-de-GQA.docx|||1NrY7N_tj_u-cK3iwpjIzjraez005KrzW
00_governanca|||CHK-Auditoria Organizacional-Timeware_junho2026.xls|||1dLEs2NkG672GBNtPByTx6_RSOLUWkkmJ
01_apoio|||PRO-GCO-001_Processo-de-Gerencia-de-Configuracao.docx|||1DjbxKyj6bxsfBJkA-tiicqTASjH0DMNN
01_apoio|||PRO-MED-001_Processo-de-Medicao.docx|||1zqYLBqG8Tjj9k-oKjR6HxyhS9f7L8KHG
01_apoio|||PRO-AQU-001_Processo-de-Aquisicao.docx|||18vVN8pTYAeP69xbHsnIGPgNLHq_IIwWt
01_apoio|||PRO-GDE-001_Processo-de-Gerencia-de-Decisoes.docx|||1N_0Tta90F4vg9JbjPvLPshqAwPCgVhxj
01_apoio|||REG-MED-001_Repositorio-Organizacional-de-Medicao.docx|||1Zbskc4JksK9Unm6Fn46DkvKMs_qmjul2
01_apoio|||PLA-GCO-001_Plano-de-Gerencia-de-Configuracao.docx|||1ZNsiKFYVP1IvPlpovyJhhKNfVfPMlibj
01_apoio|||PLA-MED-001_Plano-de-Medicao.docx|||1bl5-_poxJigRv4_5gQj8K3dFDIEqQF5e
01_apoio|||GUIA-GCO-001_Guia-de-Nomenclaturas-Tecnicas.docx|||1aSfVOk1WBb8itN0-9jMHO5yhYxsrGM2_
01_apoio|||PLA-CAP-001_Plano-de-Capacitacao.docx|||1fRYqINyOJrh7ily2FW-e2XnQH0z-mYd6
01_apoio|||PRO-CAP-001_Processo-de-Capacitacao.docx|||1QrUMeBMUAYoppls0b02CNlbDE3l6lW1-
01_apoio|||DIAG-GPC-001_Fluxo-do-Processo-Padrao.svg|||1lMaZ2M1qKHfkGfIVAhc4rRyD26-LJJPI
01_apoio/cap|||REL-CAP-001_Relatorio-de-Eficacia-2025.docx|||1vOW1WaWQyddaHXdkJfFkmVP7o4RDzxeX
01_apoio/cap|||REG-CAP-006_Sessao-Gestao-Out2025.docx|||1Qv8t4ZXt2I3tLI9pg_HOwCekuEI4JzX5
01_apoio/cap|||REG-CAP-008_Sessao-Tecnico-Fev2026.docx|||17Dvw_Q_iCQ10w0arFOOoMakjpaqGOstI
01_apoio/cap|||REG-CAP-005_Sessao-Treinamento-Mai2026.docx|||1D8oZJ-zFZ1xCaEkheX1fr_UVX0Vlfq6g
01_apoio/cap|||REG-CAP-CV-001_Indice-de-Curriculos.docx|||1QwGtPfGUwvcmpBkFxl9R60DSuJXCiunl
01_apoio/cap|||REG-CAP-004_Sessao-Treinamento-Jan2026.docx|||1PAWKGhzVbZS4SeyDrutAj9mM-HMGxKs4
01_apoio/cap|||REG-CAP-013_Consultoria-Desenho-Processos-Jun2025.docx|||1BYatkn9c-PMuHuOVurTVVcZb3UXP8Zu7
01_apoio/cap|||REG-CAP-009_Reciclagem-Mar2026.docx|||1df4AZerHPxQz0rBtvFH3zHkMramhnvzU
01_apoio/cap|||REG-CAP-011_Workshop-Azure-APIM-Mar2026.docx|||1fXmJU_h1MhW7evekDBwhWHMvT3KpvGzN
01_apoio/cap|||TPL-CAP-001_Registro-de-Sessao-de-Treinamento.docx|||1ZRcOc9EUe_tGu93X2_CHESAL31GaR5tC
01_apoio/cap|||TPL-CAP-002_Relatorio-de-Eficacia-de-Treinamento.docx|||1IyIJpISZPSI_fDzTWieJCoN1E7E-5Cgb
01_apoio/cap|||REG-CAP-010_Onboarding-Tecnico-Jan2026.docx|||1_YQ9DpRBydaU04jOIZ5JK2ooImBDHgfZ
01_apoio/cap|||REG-CAP-012_Workshop-Automacao-Testes-Fev2026.docx|||1qAcFU5xyyEbbxrLk2XL75L-EmBsJiqsi
01_apoio/cap|||REG-CAP-007_Sessao-Tecnico-Jan2026.docx|||1hVZts-J6HjJj_U41MyHZklhTMlp9xYn5
01_apoio/cap|||REG-CAP-003_Sessao-Treinamento-Ago2025.docx|||14TYg_YEqr0RMpHXS8s7SHiAJTJhV5YoI
01_apoio/cap|||REG-CAP-001B_Sessao-Reforco-Jan2025.docx|||1MKVD45n4shRUHW_6Y9s0dAbpRqrJ1Iyj
01_apoio/cap|||MAT-CAP-016_Trilha-Tech-Lead-Arquiteto.docx|||15nIEUZFERE0EJV4QZk9xpQBwlNUZFCcF
01_apoio/cap|||MAT-CAP-017_Trilha-PO-PM.docx|||1JHBn9SgnWJf6uzC3G1GGvmPIS9fZIWPa
01_apoio/cap|||MAT-CAP-015_Trilha-RH-Pessoas.docx|||1UWNTFbXxQ_kmJqSGNOkRFeX-iSomPnSG
01_apoio/cap|||REG-CAP-001_Sessao-Treinamento-Dez2024.docx|||1d3uKsFw7u-EivN-32-gbwbb2e1R9qqvt
01_apoio/cap|||MAT-CAP-022_Trilha-Responsavel-Medicao.docx|||1SrbC85jEtMo7UKhvKff-j4hV_cn6J0Gc
01_apoio/cap|||MAT-CAP-020_Trilha-QA.docx|||1Dj-qizipNW-ltxCZWvmxHzxaJTcRtZL_
01_apoio/cap|||MAT-CAP-023_Trilha-Tecnica-Onboarding.docx|||1b74pAiJa0SgmSlw6aIgafAJsLb1FO6JT
01_apoio/cap|||MAT-CAP-021_Trilha-GCO-Baseline.docx|||1CN4W49RgtDofz8gTKGpO-XG91s1JU8NG
01_apoio/cap|||GUIA-CAP-012_MiniManual-AQU.docx|||1KRovoA8y4usjBw54k-6koFndZbEtft3p
01_apoio/cap|||REG-CAP-002_Sessao-Treinamento-Mar2025.docx|||13QC7eLlcwKHSakbvo-zzE6dHLCwAxpN3
01_apoio/cap|||REG-CAP-002B_Sessao-Reforco-Set2025.docx|||1ld5yom8HOzJ8lnJjbei2JFBf-YmI4BrJ
01_apoio/cap|||MAT-CAP-014_Trilha-Time-Melhoria-SEPG.docx|||1HaW1EHn2aAQqY9xK89ODIgI2IvxWrDyP
01_apoio/cap|||MAT-CAP-018_Trilha-Devs.docx|||1dR90Hw5yPrDYtExr8YoHJM4aEYVnXWWh
01_apoio/cap|||MAT-CAP-019_Trilha-DevOps.docx|||1kSQSizbFFaPxnmhXWmsUbvmTu5xwA6-M
01_apoio/cap|||MAT-CAP-013_Trilha-COO-Portfolio.docx|||1s2-1ARa6uoXvZ242Rr1TSykDk33AhWAf
01_apoio/cap|||AVA-CAP-004_Avaliacao-GCO-ITP.docx|||1US9vp7JZU04JucCuED_goKxGr9ZZ-9g3
01_apoio/cap|||AVA-CAP-001_Avaliacao-Processo-Padrao-Geral.docx|||13GOC8q0K9e_qbFDzlZxlgbOCTnf4dNkL
01_apoio/cap|||GUIA-CAP-002_MiniManual-REQ.docx|||1Q0pAd_F2SgWNtfEyv-QNjTzg1kks8aOn
01_apoio/cap|||GUIA-CAP-011_MiniManual-CAP.docx|||1aXxOmf0e15DbrsWn1ZaraHcPjntfbJA0
01_apoio/cap|||GUIA-CAP-004_MiniManual-VV.docx|||1aVOG1oejstTXFMvTGK8o3gSs9xZjLrnB
01_apoio/cap|||GUIA-CAP-006_MiniManual-ITP.docx|||1tqFdOOTK77v62duzBsmvNVwcIigjTPvG
01_apoio/cap|||AVA-CAP-003_Avaliacao-Tecnico-REQ-PCP-VV.docx|||1TCtnSs7NzilCXO2KQAs1VUIkJGhMTACn
01_apoio/cap|||GUIA-CAP-003_MiniManual-PCP.docx|||1veYVFsie7Lr-JwZtv9R7XPIgC5G3fUgj
01_apoio/cap|||AVA-CAP-005_Avaliacao-GPC-MED-CAP.docx|||18zWHf9BsL7jy-klnk-aRjv7NYjIwYyvX
01_apoio/cap|||AVA-CAP-002_Avaliacao-GPR.docx|||1wNPYzKfOdRuIbvvDjr-otCb8GoRhCG7q
01_apoio/cap|||GUIA-CAP-001_MiniManual-GPR.docx|||1v97Th4VFwmg8-yIKfjJ5XnM8xQZ9gQ6S
01_apoio/cap|||GUIA-CAP-008_MiniManual-MED.docx|||1Y_YmTBtBwJjLQUKsAmH5ZvaVisALAAoI
01_apoio/cap|||GUIA-CAP-005_MiniManual-GCO.docx|||1e9ikwsWjrezS3sra9hJLorItDERWDv7K
01_apoio/cap|||GUIA-CAP-009_MiniManual-GPC.docx|||1SDGOr0CiUxIjNDBuODC6wdgpp8NSoeGj
01_apoio/cap|||GUIA-CAP-007_MiniManual-GDE.docx|||1prpCXsG_qqc7gT4T0IfgJGYcohNXRDl1
01_apoio/cap|||GUIA-CAP-010_MiniManual-OSW.docx|||1NX19Hv07nAVcyyvkPPEyYyUGJkKDsAWb
01_apoio/cap|||TLP-CAP-003_Formulario-de-Avaliacao-do-Instrutor.docx|||1JWxBzA34W7LlDEoejzTlBRwAGwvbR0H0
01_apoio/cap/curriculos|||CV-Flavio-Fernandes_GQA-Arquitetura.pdf|||1GSQQ2yS_E7-HnE2BFPsf1A-hkq2AVjPd
01_apoio/cap/curriculos|||CV-Cezar-Velazquez_TechLead.pdf|||1Wj_kGQvhXViFApjmjJDfvzUvbSuyqqV3
01_apoio/cap/curriculos|||CV-Caroline-Jenifer_QA.pdf|||1C7vxxTRerQM1J7J344iub0tfSV86ozhc
01_apoio/cap/curriculos|||CV-Karen-Wada_Consultora-Processos.pdf|||1jX6ax9IDMALDHpoTCrcMMerF_nGIGpCr
01_apoio/cap/curriculos|||CV-Silvio-Baroni_SEPG-GPC.pdf|||1rC4NvQveAUCAWQOkNiU8VvlVX2AkTN6i
01_apoio/cap/curriculos|||CV-Wilson-Yamada_Sponsor-Portfolio.pdf|||11gRQH_6m7Eq1ojWkfu6hRK_AqTqY65ak
02_projeto|||PRO-REQ-001_Processo-de-Engenharia-de-Requisitos.docx|||1ym-pv-GPmpf_t1ir19bOtfIgJZyX0BPt
02_projeto|||GUIA-GPR-001_Roteiro-de-Kickoff.docx|||1Iqum1Ej_iNdiaXLJ3I1sp0JrI0iWUvnk
02_projeto|||PRO-VV-001_Processo-de-Verificacao-e-Validacao.docx|||18yE_PzywceT3Z60vRqPC-yJjw4Su3528
02_projeto|||PRO-GPR-001_Processo-de-Gerencia-de-Projetos.docx|||1EWyaanDh-iLbsEOhoz5E62LDSuGvmTij
02_projeto|||PRO-PCP-001_Processo-de-Projeto-e-Construcao-do-Produto.docx|||1s0FDXL-yA-t3Q9yzDXdUt_0a8xDtib7F
02_projeto|||PRO-ITP-001_Processo-de-Integracao-do-Produto.docx|||11NtIkxH-4lYyR7a53UTbDIMAfxzzNmL6
03_templates|||TPL-GPR-005_Template-Relatorio-de-Acompanhamento.docx|||1RrYtWXZU-EYDEKO8RRN7cWL8eDs7eJBr
03_templates|||TPL-ORG-001_Template-Ata-de-Reuniao.docx|||1TcB0DCLd2cfon4P5-t6SmMDQ2H8DiNLL
03_templates|||TPL-PCP-001_Template-Documento-de-Design.docx|||1RHOuxY9-cS4dat55nWo3YnowBN9UP867
03_templates|||TPL-VV-002_Template-Registro-de-Revisao-por-Pares.docx|||1CQTzVoOCaj1aSZ3l16gqSykf09IwYRtj
03_templates|||TPL-REQ-002_Template-Matriz-de-Rastreabilidade.docx|||1wRjUaxEY4zg7_yfZTTfB01iEETZiYN5T
03_templates|||TPL-REQ-001_Template-Documento-de-Requisitos.docx|||1CjKEf7sYu272riKdQQ5L4YbyfM4KqNe0
03_templates|||TPL-VV-001_Template-Plano-de-Verificacao-e-Validacao.docx|||11M6t6WkYgji_gubuFgyKj7WmRvLYnnYI
03_templates|||TPL-GPR-006_Template-Change-Request.docx|||1eA_EZaEl4h-zUvvmULGvgDT-f_zzSARu
03_templates|||TPL-GPC-001_Template-Registro-de-Verificacao-de-GQA.docx|||19rnaNwNh_stfJKjPUPjCKs5Yl6USfbxY
03_templates|||TPL-ITP-001_Template-Estrategia-de-Integracao.docx|||1mj5YGF4Ph8FMqz3DRx8ti-vrPmrFZiAb
03_templates|||TPL-GPR-004_Template-Termo-de-Encerramento.docx|||1NgZq_uG9A-W3qCyb89O4VDZBpJf5QeKR
03_templates|||TPL-GPR-003_Template-Registro-de-Adaptacao.docx|||1oKWodTuwtumhMDYy6eGs2cIuaJDT828w
03_templates|||TPL-GPR-001_Template-Plano-de-Projeto.docx|||190E2DG0B4jgwqg4oJ8uTXDUNNd1ROXkM
03_templates|||TPL-GPR-002_Template-Termo-de-Abertura.docx|||1dFy9pqLoURReKkf88DW9zlI5isnqV9BG
03_templates|||TPL-GDE-001_Template-Registro-de-Analise-de-Decisao.docx|||1wHbMGhvGpwUGtEj4oPQQ1KV97zLwEy7n
03_templates|||EXEMPLO-GPR-005_Status-Report-Exemplo-Preenchido.docx|||1L0rob_f1ODSOZXHoXgV_w_znX6JwKNga
04_registros/AASP_CNJ|||GEST-AASPCNJ01-001_Planilha-de-Gestao-do-Projeto.xlsx|||1Vdng-5mBCjlXpajdZ6ScLD67c8e85qob
04_registros/AASP_CNJ|||00_INDICEAASPCNJ01_MapadeRegistros.docx|||1kf3EdS_AsVImdD6P_03zCt-OF7_UHj6N
04_registros/AASP_CNJ|||CAP-AASPCNJ01-001_Registro-de-Capacitacao-da-Equipe.docx|||11db2exVuX-huVhgL8azaDzByfMFQv_pb
04_registros/AASP_CNJ|||RAC-AASPCNJ01-001_Relatorio-de-Acompanhamento.docx|||1G_0Odyq4HbI3nfCyDh7-zZoXDgu_BiON
04_registros/AASP_CNJ|||GCO-AASPCNJ01-001_Registro-de-Configuracao.docx|||1IJozbK4BobNpyION4zPWa_s2auynxhv9
04_registros/AASP_CNJ|||GQA-AASPCNJ01-001_Registro-de-GQA.docx|||1AI6TXKQ5f6Oh5g6VmFgzTmXNQHEtLOlh
04_registros/AASP_CNJ|||RASTR-AASPCNJ01-001_Matriz-de-Rastreabilidade.docx|||1lfPyboZZ3hd-8g3IzjQRQciVYXCztgZw
04_registros/AASP_CNJ|||REL-VV-AASPCNJ01-001_Relatorio-de-Execucao-de-Testes.docx|||1whC8gTS2Mz1CW7QajDlOMvfXNlxt5AP6
04_registros/AASP_CNJ|||PLA-AASPCNJ01-001_Plano-de-Projeto.docx|||1y9zo8Fx1AewE2z9GVdmhsecKySyDoSf6
04_registros/AASP_CNJ|||VV-AASPCNJ01-001_Plano-de-VV.docx|||199h316ZoZ43k7W_rnPpNq6GTBr502zuu
04_registros/AASP_CNJ|||CR-AASPCNJ01-001_Change-Request.docx|||1J3XOn56R7ruWa8vhPAnVzI7r092DFLY-
04_registros/AASP_CNJ|||PCP-AASPCNJ01-001_Documento-de-Design.docx|||1ywDtDrM3zHptY2XfL6ut22ra3MEbuLTe
04_registros/AASP_CNJ|||ITP-AASPCNJ01-001_Estrategia-de-Integracao.docx|||1tobpnNoEQf5LarNnPEJAEJnt5BAZGrtQ
04_registros/AASP_CNJ|||REV-AASPCNJ01-001_Registro-de-Revisao-por-Pares.docx|||1PfM9kKAjFJofhtJn5xqgicqG-0-v3mA-
04_registros/AASP_CNJ|||GDE-AASPCNJ01-001_Registro-de-Analise-de-Decisao.docx|||1EpobYWmsjCftVdxosshioK0mCuXVFlEq
04_registros/AASP_CNJ|||ADAP-AASPCNJ01-001_Registro-de-Adaptacao.docx|||14GFOTJj6jxDsV62h60zCqIgxnRW-DbWW
04_registros/AASP_CNJ|||MED-AASPCNJ01-001_Registro-de-Medicao.docx|||1ePiJ3ONIXZf2gN00FzXuyUCIXdrpezsx
04_registros/AASP_CNJ|||ATA-AASPCNJ01-001_Ata-Alinhamento-Fluxo-CNJ.docx|||12sOcC6FDaP-3OdoM1_duAMNNcbgYsQvb
04_registros/AASP_CNJ|||TAP-AASPCNJ01-001_Termo-de-Abertura.docx|||1hKBUW2dTTfmd8PEs36tEuIQjsKjPAuRF
04_registros/AASP_CNJ|||REQ-AASPCNJ01-001_Documento-de-Requisitos.docx|||1GDwHhtK1Nc-2AptWo8HIryVHeOF8hzj5
04_registros/AASP_Andamento-Processuais|||00_INDICE-AASPAP01_Mapa-de-Registros.docx|||1q7KCwgqQ5w0lUx4x26DYYFCx5gMmuglM
04_registros/AASP_Andamento-Processuais|||CR-AASPAP01-002_Change-Request.docx|||1e495VArI_U_3TMW7KQc0p3GfqUiAek4a
04_registros/AASP_Andamento-Processuais|||GEST-AASPAP01_Planilha-de-Gestao-do-Projeto.xlsx|||1ZB2yOn7Pt1eviyaL1-XMCaiamiCql6gz
04_registros/AASP_Andamento-Processuais|||CR-AASPAP01-001_Change-Request.docx|||1p94Xtd_QSXghPgdDQiaj9m8JoArrHree
04_registros/AASP_Andamento-Processuais|||REV-AASPAP01-001_Registro-de-Revisao-por-Pares.docx|||1b2XcTr_rrP1ntnEzllbB2gHNdFaEaLsQ
04_registros/AASP_Andamento-Processuais|||TAP-AASPAP01-001_Termo-de-Abertura.docx|||1vj9arb9bNdy4z2DLV7U8sEbBp4GuIMHw
04_registros/AASP_Andamento-Processuais|||VV-AASPAP01-001_Plano-de-VV.docx|||11Vx-bMcQfovROo3RwVY7-dQClEVcFp1e
04_registros/AASP_Andamento-Processuais|||GCO-AASPAP01-001_Registro-de-Configuracao.docx|||1Rl6GXnN1H9mRl0V6bSYw9Pl_D-0THH_O
04_registros/AASP_Andamento-Processuais|||MED-AASPAP01-001_Registro-de-Medicao.docx|||17jDdprZJdXPr-muQAYnQNXoMNU1VOSAV
04_registros/AASP_Andamento-Processuais|||PLA-AASPAP01-001_Plano-de-Projeto.docx|||1D7tchEQKaqeZMZ2vNNa0wgnXevI-Br8H
04_registros/AASP_Andamento-Processuais|||ITP-AASPAP01-001_Estrategia-de-Integracao.docx|||1cSlw8Kqi3fPboNG2KDMoRuWYLii5s1Qm
04_registros/AASP_Andamento-Processuais|||ADAP-AASPAP01-001_Registro-de-Adaptacao.docx|||1K4Q7uEZFdHy_r6nJhALzqCO-aH9fSg0o
04_registros/AASP_Andamento-Processuais|||RAC-AASPAP01-001_Relatorio-de-Acompanhamento.docx|||155bBfagbpaZdKxRSqXQzkmZSQkmCHNt0
04_registros/AASP_Andamento-Processuais|||RASTR-AASPAP01-001_Matriz-de-Rastreabilidade.docx|||1nqVdGrDD5CJ_3zPB3IehiBEhuDKSxZlS
04_registros/AASP_Andamento-Processuais|||ATA-AASPAP01-001_Ata-Alinhamento-Fluxo-CNJ.docx|||1xnhEepqAaUxgcarVXwtyBITSBo4LiLVq
04_registros/AASP_Andamento-Processuais|||REL-VV-AASPAP01-001_Relatorio-de-Execucao-de-Testes.docx|||1TzIWFlAfEpqwWO0JIF3DACKPzPgC2sPV
04_registros/AASP_Andamento-Processuais|||REQ-AASPAP01-001_Documento-de-Requisitos.docx|||1kV_N6w7TU5Cwe5SYWL3tcptlcr4oazYt
04_registros/AASP_Andamento-Processuais|||PCP-AASPAP01-001_Documento-de-Design.docx|||1MkxkyhMa5O4d9ee1JVYHNZV59BxiHU3m
04_registros/AASP_Andamento-Processuais|||GDE-AASPAP01-001_Registro-de-Analise-de-Decisao.docx|||1sDGfP0m6WdyKS4MYzHLLZNNQCyus0xJw
04_registros/AASP_Andamento-Processuais|||GQA-AASPAP01-001_Registro-de-GQA.docx|||19WE56UNWv-aVGiPL24PzStXvX1NuplXF
04_registros/AASP_Andamento-Processuais/evidencias|||andamentos_historicoPreservado2.jpg|||1l8sW0voE8k-y6ew1dTbd5aePqlr5UffM
04_registros/AASP_Andamento-Processuais/evidencias|||andamentos_ApisParceiras.jpg|||1di53G9P1RLs2VtusPoTRJay3fOG7c1L2
04_registros/AASP_Andamento-Processuais/evidencias|||devops_andamentos_commits.png|||1fkX4bzm8iFGVFOtMtIK4rEtjZhF1bomY
04_registros/AASP_Andamento-Processuais/evidencias|||andamentos_historicoPreservado1.jpg|||1zXIyrAtYHinI0oETDvC-PQ0sgF4Z6Fjy
04_registros/AASP_Andamento-Processuais/evidencias|||devops_andamentos_branches.png|||1YaLDEwp8l6FoyFTaMTEWMmeLqbT6gfnx
04_registros/AASP_Andamento-Processuais/evidencias|||devops_andamentos_pr_detail_2.png|||1oBhgj_wDBy_Ap5QoTy-HRRu7a62wsa6S
04_registros/AASP_Andamento-Processuais/evidencias|||devops_andamentos_prs.png|||14Se4mDLpHgV5EjT8w7kd1TypMmOEmzMU
04_registros/AASP_Andamento-Processuais/evidencias|||devops_andamentos_estrutura.png|||1w_Hj4o0YyA0oYRSE1_5d1VE67_JisEFY
04_registros/AASP_Andamento-Processuais/evidencias|||andamentos_historicoPreservado.jpeg|||1Nx8OtLH1th3lVs3pRHd1yJjZ-l5jsR4s
04_registros/AASP_Andamento-Processuais/evidencias|||devops_andamentos_pipeline_ci.png|||1p1VezDp1C0XYPSOjQk978cQmrvGVrTWo
04_registros/AASP_Andamento-Processuais/evidencias|||andamentos_arquitetura.svg|||1hwULBFas2YSOqC3BIi0dVbGelx-U-Mp0
04_registros/AASP_Andamento-Processuais/evidencias|||devops_andamentos_pipeline_cicd.png|||1Zf1k0e36NnyOxq-zv6Rrz22N5tvM8x49
04_registros/AASP_Andamento-Processuais/evidencias|||devops_andamentos_tags.png|||1fGjCYwKyDczuhAJPkpaxCMdrrCi7BRc4
04_registros/AASP_Andamento-Processuais/evidencias|||andamentos_retry_webhook.jpeg|||15lXDLmdWW620DS2qE9-qwIEzvAgmpGlW
04_registros/AASP_Andamento-Processuais/evidencias|||estrutura_andamentosProcessuais.png|||1BQ_7ZiscsMUDn84a_vBXcghhchm29hLs
04_registros/AASP_Andamento-Processuais/evidencias|||devops_andamentos_pr_detail_1.png|||1FSv2kmZbKNXOk5tjctohUFqHtn-BZ9bu
04_registros/PROFARMA_Cadastro-de-Clientes|||ATA-PROFARMA01-002_Ata-de-Aceite-Final.docx|||1odIVE62I5L7iqqLFUCACbHt3w2HRB8yR
04_registros/PROFARMA_Cadastro-de-Clientes|||PLA-PROFARMA01-001_Plano-de-Projeto.docx|||189Kp-xA8BJ3SsO9NhxDyDOUtX3dADifK
04_registros/PROFARMA_Cadastro-de-Clientes|||ADAP-PROFARMA01-001_Registro-de-Adaptacao.docx|||1aRgkciq4Ed3iPJgEbe1jJnpkXKlF4Tma
04_registros/PROFARMA_Cadastro-de-Clientes|||RAC-PROFARMA01-001_Relatorio-de-Acompanhamento.docx|||1HjCxyFwvC7Fxp4OXwxjqqL7OyYlSfI9x
04_registros/PROFARMA_Cadastro-de-Clientes|||TAE-PROFARMA01-001_Termo-de-Encerramento.docx|||1k_UZ_jWRVmpHA1Iiysa7vKvUwQ1T8QED
04_registros/PROFARMA_Cadastro-de-Clientes|||ITP-PROFARMA01-001_Estrategia-de-Integracao.docx|||1DKKcqAZK78dkV2UIRRSXHtELjnxdw4aj
04_registros/PROFARMA_Cadastro-de-Clientes|||GEST-PROFARMA01_Planilha-de-Gestao-do-Projeto.xlsx|||12SEFOVCAA1jVQRgM1c1RDR_CL3KaW32_
04_registros/PROFARMA_Cadastro-de-Clientes|||TAP-PROFARMA01-001_Termo-de-Abertura.docx|||13BKbRnLmlBDMBkioH54vOpmicKqlei8W
04_registros/PROFARMA_Cadastro-de-Clientes|||LI-PROFARMA01-001_Licoes-Aprendidas.docx|||1NMayhfr_e6Li-R0NEGoST5QATJgifMc7
04_registros/PROFARMA_Cadastro-de-Clientes|||REQ-PROFARMA01-001_Documento-de-Requisitos.docx|||17OE2PW4XkEy3OnJlSb6rtsHHh9ouTFJE
04_registros/PROFARMA_Cadastro-de-Clientes|||MED-PROFARMA01-001_Registro-de-Medicao.docx|||1nmbiWfvYMAAa7Y5pO14VxaZgxEPQuKOn
04_registros/PROFARMA_Cadastro-de-Clientes|||CR-PROFARMA01-001_Registro-de-Change-Requests.docx|||1XWRRoeJ6NnvjCG3S0hlyPyxnMfkf9vM7
04_registros/PROFARMA_Cadastro-de-Clientes|||GCO-PROFARMA01-001_Registro-de-Gerencia-de-Configuracao.docx|||1ld1PXP0yafu0eW7nUNXN0iRCm1P7cskc
04_registros/PROFARMA_Cadastro-de-Clientes|||RASTR-PROFARMA01-001_Matriz-de-Rastreabilidade.docx|||1ep0JvVMNhdLVRwWXVMIyfLxurd7ReAyI
04_registros/PROFARMA_Cadastro-de-Clientes|||GQA-PROFARMA01-001_Registro-de-GQA.docx|||1hefq-kmdiYVOIvc0RoWHcSVQDBbL8Fq4
04_registros/PROFARMA_Cadastro-de-Clientes|||ATA-PROFARMA01-001_Ata-de-Kickoff.docx|||1uNkVDge5WAgwLTCcUFZmBLlqKog16nmG
04_registros/PROFARMA_Cadastro-de-Clientes|||REV-PROFARMA01-001_Registro-de-Revisao-Tecnica.docx|||1dHd-EpgVX2Mg2Wo5Y_uxB8KFH4qphhxM
04_registros/PROFARMA_Cadastro-de-Clientes|||PCP-PROFARMA01-001_Documento-de-Design.docx|||1t0fD748MR_EBRMHsgQzVGLKRnJ4Zq7__
04_registros/PROFARMA_Cadastro-de-Clientes|||CTQ-PROFARMA01-001_Cenarios-de-Teste-Homologacao.docx|||1ts70FWbIasG3sf9qDBSw8cXldJfRJ7bY
04_registros/PROFARMA_Cadastro-de-Clientes|||VV-PROFARMA01-001_Plano-de-VV.docx|||15tRkpsMad47tD4gPdlLKImY5ipIadjxw
04_registros/PROFARMA_Cadastro-de-Clientes|||REL-VV-PROFARMA01-001_Relatorio-de-Execucao-de-Testes.docx|||1KF8zj0lufb0VPg6riJwGhibojJJIopHM
04_registros/PROFARMA_Cadastro-de-Clientes|||GDE-PROFARMA01-001_Registro-de-Analise-de-Decisao.docx|||1wU_hxVhwrX-XUT1OQmEKxw80Mz0T5Rat
04_registros/PROFARMA_Cadastro-de-Clientes|||Analise-Decisao_PROFARMA_TIMEWARE.xlsx|||1iS1LRiviBLZgmxymc7Hfq7szzjUGkHjK
04_registros/PROFARMA_Cadastro-de-Clientes|||EV-Jira-PROFARMA01-001.docx|||1ftGwgagbTkMohz_ZpKcLT0Z1_cPq_4A9
04_registros/PROFARMA_Cadastro-de-Clientes|||EV-Swagger-PROFARMA01-001.docx|||17hu5hEzQctuc6OaDun7ki1pn1f-fp7cL
04_registros/PROFARMA_Cadastro-de-Clientes|||EV-AzureDevOps-PROFARMA01-001.docx|||1K3LjcxoNE5lRsuUe0IIkCtwreoj3Bizv
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||PLA-GASMIG02-001_Plano-de-Projeto.docx|||1ZIwX-h0RDNJfhO5AXev2vqhYrGHOr2qf
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||GEST-GASMIG02_Planilha-de-Gestao-do-Projeto.xlsx|||1r_dSFP9AmEnj79FkmqU80OkuYWnwzfZS
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||ADAP-GASMIG02-002_Registro-de-Adaptacao-OS002.docx|||1oQshhvn5DIq1DoB3rybc8jEbP_OdEnfn
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||ADAP-GASMIG02-001_Registro-de-Adaptacao.docx|||1uDnjyU3hDOdzbVZPUFA-1PGqXCsuDeS7
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||TAP-GASMIG02-002_Termo-de-Abertura-OS002.docx|||1cMUOhgzZzj6wwWIijFMDugRVZIl8aPvf
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||PLA-GASMIG02-002_Plano-de-Projeto-OS002.docx|||1jfUDk7bhCtOXUYDcBnpjOx8FARp2VqPH
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||LI-GASMIG02-001_Licoes-Aprendidas.docx|||1ra9yyJuVYa2tmQW0DKwbJgannxAI2W0E
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||ATA-GASMIG02-003_Apresentacao-Entrega-OS002.docx|||1YCW3PFHS2p1bh8O6VWyZ8y2NFoG6ZWFN
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||VV-GASMIG02-001_Plano-de-VV.docx|||15t1yh5oMbpFZ47rREDx9kJPgS8oXyRmy
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||MED-GASMIG02-001_Registro-de-Medicao.docx|||1LtjN_WSjn_g1wQ_RRLe4Z4KP6Cnu-Ohx
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||RAC-GASMIG02-001_Relatorio-de-Acompanhamento.docx|||1bAl8o8UPkTb1AHypGNvRxEJG8UYswpvS
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||GQA-GASMIG02-001_Registro-de-GQA.docx|||1V-T1g_9AKu9TLFPROMR0fA1gkyx180a-
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||VV-GASMIG02-002_Plano-de-VV-OS002.docx|||18GgauiJykVy8AASSKtJcqkt2FR-6AuRv
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||TAE-GASMIG02-002_Termo-de-Encerramento-OS002.docx|||1-zvT45vPsVTAwmcEzQVak8qKT4dnQWEr
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||GCO-GASMIG02-001_Gerencia-de-Configuracao.docx|||1qpIzDVbsmVYkiXW9NNdciOuv5iovMNu4
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||ITP-GASMIG02-002_Estrategia-de-Integracao-OS002.docx|||1JIAbQHVpTSO3V0AKGnWyVwEs6zuTWX7R
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||RASTR-GASMIG02-001_Matriz-de-Rastreabilidade.docx|||1_delTbo-PEeRpuGGGagv5p2SOF-ZNRos
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||RASTR-GASMIG02-002_Matriz-de-Rastreabilidade-OS002.docx|||1tfy9oNEOsThCKNb1YrvQq8Ll1jX9kYJO
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||REQ-GASMIG02-001_Documento-de-Requisitos.docx|||13duq0igfGKvJAHKh_sD9TN7WqTpuOfdn
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||REV-GASMIG02-001_Registro-de-Verificacao-Tecnica.docx|||1biHfNjSf_g7hnUcjBATCG48aG7ARNvRL
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||ATA-GASMIG02-001_Ata-de-Kickoff.docx|||1XcstqfQTQC0cm8wZgfc9oLgjSknU6Ueq
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||ATA-GASMIG02-002_Ata-de-Aceite-OS001.docx|||1CcRbEbIcDWeHO7sOznH_FcuK-uVVusM0
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||PCP-GASMIG02-001_Documento-de-Design.docx|||1YI-40O6IWGMB0C_R6ytpbxc2w1E7W4fF
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||TAE-GASMIG02-001_Termo-de-Encerramento-OS001.docx|||1mw7eJBj6HANyIYJ1MgxEnO59ML9sl5aa
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||TAP-GASMIG02-001_Termo-de-Abertura.docx|||1WW3VNRafS44JkZiXK29Rm7bYI7NbP9rS
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||CAP-GASMIG02-001_Registro-de-Capacitacao-da-Equipe.docx|||1ibleg7tRQDHfYp3ahRpAlbYjKetFDzzT
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||REQ-GASMIG02-002_Documento-de-Requisitos-OS002.docx|||1_yN7yHPBc9iKilAlxd4pxj52G1_WoCHP
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||GDE-GASMIG02-001_Registro-de-Analise-de-Decisao.docx|||1sqAtX-R1-EjASMtMCnp_fjiessTaSaQU
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||PCP-GASMIG02-002_Documento-de-Design-OS002.docx|||1Q4M2R4uqiy81BP8L3HiQ_idRqmpMt8Cw
04_registros/FTGASMIG_Governanca-APIs/GASMIG02_Fundacao-Tecnologica|||AnaliseDecisao_FTGASMIG_TIMEWARE.xlsx|||1u5oCqJrnQuaPBUMqTNQvaukwxqZvsmQi
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||PLA-FRUKI01-001_Plano-de-Projeto.docx|||1WEPaZNg7rxj29B48l_scz7x7CyLlvoSf
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||GEST-FRUKI01-001_Planilha-de-Gestao-do-Projeto.xlsx|||1RsktvVNbrH1f-RmcV3_AmIu3uJt9Q7sH
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||ATAFRUKI01008_AtaKickoffPacoteFinal24.docx|||11alVpwlnsPjgd1_PO4OzMBXeJLRz6PXi
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||LI-FRUKI01-001_Licoes-Aprendidas.docx|||1pXvWCErdbUdvvAcXV93935DeGW6bJ7xL
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||ADAP-FRUKI01-002_Registro-de-Adaptacao-PacoteFinal24.docx|||1LNzyoQkTM7oLgWtSCPY6VxNqt4JS_MzM
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||ADAP-FRUKI01-001_Registro-de-Adaptacao.docx|||1NaEJlUDHw0FE6fgH28gpifHft7ZLxUyL
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||TAP-FRUKI01-002_Termo-de-Abertura-PacoteFinal24.docx|||1UUmn_lYKBv8eG5UZBaLipFbSMdnYTxgU
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||TAE-FRUKI01-002_Termo-de-Encerramento-PacoteFinal24.docx|||1D3T0RHrOPhOq08Edxc1fWW-cZ1uR-e2n
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||MED-FRUKI01-001_Registro-de-Medicao.docx|||1D6gCK_SBP31WdmYGToQH2iCZV1RdzhWI
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||RASTR-FRUKI01-002_Matriz-de-Rastreabilidade-PacoteFinal24.docx|||1S4nR_PkZnGQrm6ppWv47ftgH2ptyG9xJ
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||TAE-FRUKI01-001_Termo-de-Encerramento.docx|||1JNvLtWIaqno7WXqzPfQwcK1bRE3UjCRv
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||ATA-FRUKI01-002_Ata-Levantamento-Metas.docx|||1cypxWHL-NU9CPnLV47N-bKnP0QrSIB17
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||ITP-FRUKI01-001_Estrategia-de-Integracao.docx|||19YCw05XWIRh-NaI5GWf1c-MF6m5AhIra
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||GCO-FRUKI01-001_Registro-de-Configuracao.docx|||1fMcui7ebx_4af41JtRltlCCJzDWkd1eA
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||ATA-FRUKI01-005_Ata-Validacao-Sprint2-RegraDeOuro.docx|||1VmGwsZK4GLu_zpzmrHnD2XOChsiR1GZr
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||REQ-FRUKI01-002_Documento-de-Requisitos-PacoteFinal24.docx|||1GexOXykQE4rnKFPmeW1kVQac18vsnlnQ
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||VV-FRUKI01-001_Plano-VeV.docx|||1yCFY98oM8rlf44HUIeERib4_7jtYsKeU
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||PLA-FRUKI01-002_Plano-de-Projeto-PacoteFinal24.docx|||1d0l6IhtfbI9hE0UYxQm6nVq9grpjKnpn
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||ATA-FRUKI01-001_Ata-de-Kickoff.docx|||1pjAXblbrxR8Y6p5eq2BDYRBpiYlK3Ag_
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||ITP-FRUKI01-002_Estrategia-de-Integracao-PacoteFinal24.docx|||1A-6S0ZZmHteTzIjSNP46H2OJYgeI-Gc7
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||PCP-FRUKI01-001_Documento-de-Design.docx|||1ff6Ha1FH1T1XAJGnmdy2QvkrqXxCUg2s
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||CR-FRUKI01-001_Solicitacao-de-Mudanca-RegraDeOuro.docx|||1z2TIuyQQTiU7fOKmK0kQfziTtE8NqiLc
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||GQA-FRUKI01-001_Registro-de-GQA.docx|||1jmdev8ggRDsbM_CrkMQY_oKaImvKG4vM
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||ATA-FRUKI01-003_Ata-de-Aceite-Final.docx|||1m0odZUeHy212r9AQwyZ_sfwXzznT4oqJ
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||REQ-FRUKI01-001_Documento-de-Requisitos.docx|||1mCQ9GroaEFEjulblg31P0Kdknw95w8LJ
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||RAC-FRUKI01-001_Relatorio-de-Acompanhamento.docx|||1_C3hZUp-7-2ldAa2BmIfZ3tMfa9WX0bd
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||ATA-FRUKI01-007_Ata-Piloto-Pacote1.docx|||1bNAl7ljD0PhTyZEb-2ngdIbBzCpvNY2q
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||VV-FRUKI01-002_Plano-VeV-PacoteFinal24.docx|||1d_okleyBKoPtPW2zEy7300KiXoWL62oj
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||ATA-FRUKI01-004_Ata-Validacao-Sprint1-NaoAlocados.docx|||10Qo6yB8nqAQmzacMbuLZaVv1ZaX178Lt
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||PCP-FRUKI01-002_Documento-de-Design-PacoteFinal24.docx|||17KodAvypysQkbqocfQIrlE2q1dfRIUnY
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||RASTR-FRUKI01-001_Matriz-de-Rastreabilidade.docx|||1SokajsndvgCtPAX7FyUpNp9mIDxQrVQf
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||TAP-FRUKI01-001_Termo-de-Abertura.docx|||1ehO_9gQoPrLS-Ykhxfnnsrn9KZ9o8zdH
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||GDE-FRUKI01-001_Registro-de-Decisao.docx|||1h8FKsNP59PnO8tzm2WEJ_vgsZYRSmjwX
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||ATA-FRUKI01-006_Ata-Validacao-Sprint3-PDV.docx|||1ujgsYTHnAVE_J60APn-OqGSRaCTbfPHf
04_registros/FTFRUKI_SuperApp-Forca-de-Vendas|||AnaliseDecisao_FTFRUKI_TIMEWARE.xlsx|||19sYcZ40SHiEe7qUV7jau-3fruYrAzjWp
05_capacidade|||MAPA-ORG-001_Matriz-de-Papeis-e-Responsabilidades.docx|||1j98CQbhiUr74uV7VenwJcEhnLmLupxhe
05_capacidade|||MAPA-CAP-001_Mapa-de-Capacidade-dos-Processos.docx|||17csazhVOlhhgaMN-DCmoQEUiJsfrN1Il
05_capacidade|||MAPA-ORG-001_Matriz-de-Papeis-e-Responsabilidades.xlsx|||1i87bnTiddeMILoHq3CBlyZ-WzPS5meEW
"""

def main():
    out_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    rows = []
    seen_ids = {}
    seen_pathname = {}
    dup_ids = []
    dup_names = []
    for raw in DATA.strip().splitlines():
        raw = raw.strip()
        if not raw:
            continue
        pasta, nome, fid = raw.split("|||")
        ext = nome.rsplit(".", 1)[-1].lower() if "." in nome else ""
        link = LINK_TPL.format(id=fid)
        rows.append((pasta, nome, ext, link, fid))
        if fid in seen_ids:
            dup_ids.append((fid, nome, seen_ids[fid]))
        else:
            seen_ids[fid] = nome
        key = (pasta, nome)
        if key in seen_pathname:
            dup_names.append(key)
        else:
            seen_pathname[key] = True

    # ordena por pasta e depois por nome do documento
    rows.sort(key=lambda r: (r[0], r[1]))

    csv_path = os.path.join(out_dir, "MAPA-DRIVE_Indice-de-Links.csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Pasta", "Documento", "Tipo", "Link"])
        for pasta, nome, ext, link, fid in rows:
            w.writerow([pasta, nome, ext, link])

    # tenta gerar xlsx
    xlsx_ok = False
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        wb = Workbook()
        ws = wb.active
        ws.title = "Mapa de Links Drive"
        ws.append(["Pasta", "Documento", "Tipo", "Link"])
        for c in ws[1]:
            c.font = Font(bold=True)
        for pasta, nome, ext, link, fid in rows:
            ws.append([pasta, nome, ext, link])
            cell = ws.cell(row=ws.max_row, column=4)
            cell.hyperlink = link
            cell.font = Font(color="0563C1", underline="single")
        widths = {"A": 60, "B": 60, "C": 8, "D": 70}
        for col, wd in widths.items():
            ws.column_dimensions[col].width = wd
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:D%d" % (len(rows) + 1)
        xlsx_path = os.path.join(out_dir, "MAPA-DRIVE_Indice-de-Links.xlsx")
        wb.save(xlsx_path)
        xlsx_ok = True
    except Exception as e:
        xlsx_path = None
        print("[aviso] xlsx nao gerado (%s)" % e, file=sys.stderr)

    # resumo
    from collections import Counter, defaultdict
    by_ext = Counter(r[2] for r in rows)
    by_folder = Counter(r[0] for r in rows)
    by_folder_docx = defaultdict(int)
    for r in rows:
        if r[2] == "docx":
            by_folder_docx[r[0]] += 1

    print("=" * 72)
    print("MAPA DE LINKS — ACERVO OFICIAL TIMEWARE (Docs oficial Timeware)")
    print("=" * 72)
    print("Total de arquivos mapeados: %d" % len(rows))
    print("IDs unicos: %d" % len(seen_ids))
    print()
    print("Por tipo:")
    for ext, n in sorted(by_ext.items(), key=lambda x: (-x[1], x[0])):
        print("  %-6s %d" % (ext or "(sem)", n))
    print()
    print("Por pasta (total / docx):")
    for folder in sorted(by_folder):
        print("  %-70s %3d / %3d docx" % (folder, by_folder[folder], by_folder_docx.get(folder, 0)))
    print()
    if dup_ids:
        print("!! IDs DUPLICADOS:")
        for fid, nome, antes in dup_ids:
            print("   %s  %s (ja visto como %s)" % (fid, nome, antes))
    else:
        print("OK: nenhum ID duplicado.")
    if dup_names:
        print("!! (pasta,nome) DUPLICADOS:")
        for k in dup_names:
            print("   %s" % (k,))
    else:
        print("OK: nenhum (pasta,documento) duplicado.")
    print()
    print("CSV : %s" % csv_path)
    if xlsx_ok:
        print("XLSX: %s" % xlsx_path)

if __name__ == "__main__":
    main()
